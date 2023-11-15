import requests
import json
from openpyxl import Workbook, load_workbook
import time

# 엑셀 파일에 데이터 추가하는 함수
def append_to_excel(data, filename='exam4you_data.xlsx'):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Category", "Title", "URL"])  # 헤더 추가

    for entry in data:
        ws.append([entry['category'], entry['title'], entry['url']])

    wb.save(filename)
    print("Data appended to Excel!")

# 세션 객체 생성
session = requests.Session()

# 요청 헤더 설정
headers = {
     "Host": "bbs.exam4you.com",
    "Cookie": "_gid=GA1.2.229133552.1696695017; JSESSIONID=3C639EF5EE3DDF81F13840E1BE49296B; AWSELB=61BF71391CCF8A0434BDD9AD0BA45C24A96E4220B0D5166FA1F91AAF7A12FC5FB02ABB99F014DA48EE4093604CE5C93CC6DB873BDEF8FDE40947B3AA6F2921E375DB4BE493; AWSELBCORS=61BF71391CCF8A0434BDD9AD0BA45C24A96E4220B0D5166FA1F91AAF7A12FC5FB02ABB99F014DA48EE4093604CE5C93CC6DB873BDEF8FDE40947B3AA6F2921E375DB4BE493; wcs_bt=90760b3f69ef88:1696695095; _gat_gtag_UA_74988519_1=1; _ga=GA1.1.552363514.1696385750; _ga_ETJKH7XM2J=GS1.1.1696695016.3.1.1696695105.46.0.0",
    "Content-Length": "166",
    "Sec-Ch-Ua": '"Not A(Brand";v="24", "Chromium";v="110"',
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Content-Type": "application/json",
    "X-Requested-With": "XMLHttpRequest",
    "Sec-Ch-Ua-Mobile": "?0",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.5481.78 Safari/537.36",
    "Sec-Ch-Ua-Platform": '"Linux"',
    "Origin": "https://bbs.exam4you.com",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Dest": "empty",
    "Referer": "https://bbs.exam4you.com/app/user/bbsShare/listPage?bbsPart=B01-04",
    "Accept-Encoding": "gzip, deflate",
    "Accept-Language": "en-US,en;q=0.9",
    "Connection": "close",
}

# 초기 페이지 번호와 URL
start_page = 1
end_page = 10
url = "https://bbs.exam4you.com/app/user/bbsShare/getList.json"

# 페이지 범위에 대한 반복
for pageNo in range(start_page, end_page + 1):
    payload = {
        "bbsPartGubun": "", "bbsPart": "B01-04", "errArea": "", "categoryName": "", "openwordName": "", "adminRegYn": "",
        "pageNo": pageNo, "regUserSeq": "0", "rowCount": "20", "totalcount": "0" if pageNo == 1 else "545"
    }
    response = session.post(url, headers=headers, data=json.dumps(payload))
    
    # 요청 사이에 1초 지연
    #time.sleep(1)

    if response.status_code == 200:
        try:
            response_data = response.json()
            formatted_data = [{
                "category": item.get("categoryName", ""),
                "title": item.get("title", ""),
                "url": f"https://bbs.exam4you.com/app/user/bbsShare/viewPage?listSeq={item.get('listSeq', '')}&noticeYn=N&bbsPartGubun=&bbsPart=B01-04&categoryName=&openwordName=&searchTypeStr=title&searchWord=&pageNo=1&countPerPage=20&refererPage=listPage&errArea=&totalCount=545&userSeq=0&ad="
            } for item in response_data.get("cList", [])]
            append_to_excel(formatted_data)
        except json.JSONDecodeError as e:
            print(f"JSON decoding failed: {e}")
            print(f"Response content: {response.text}")
    else:
        print(f"Error: HTTP {response.status_code}")
        print(response.text)
